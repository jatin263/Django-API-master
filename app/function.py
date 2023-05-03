from openpyxl import load_workbook
import os
import datetime
import speech_recognition as sr
import subprocess

def xlsxToList(s):
    file = load_workbook(s)
    sheet = file.active
    maxRow=sheet.max_row
    maxCol=sheet.max_column
    ans=[]
    for i in range(2,maxRow+1):
        a=[]
        for j in range(1,maxCol+1):
            a.append(sheet.cell(row=i,column=j).value)
        ans.append(a)
    ind=s.index('.')
    x = datetime.datetime.now()
    x=str(x)
    x=x.replace(':',"-")
    x=x.replace(".","-")
    os.rename(s,s[:ind]+x+s[ind:])
    return ans

def handle_Xlsx_file(f):
    with open('app/static/upload/xlxs/'+f.name,'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return xlsxToList('app/static/upload/xlxs/'+f.name)

def speechToText(s):
    r = sr.Recognizer()
    with sr.AudioFile(s) as source:
        audio_text = r.record(source)
        try:
            # using google speech recognition
            text = r.recognize_google(audio_text)
            print(text)
            return [text,s]
        except:
            return ['Sorry.. run again...',s]



def converterOfFile(s):
    out=s[:len(s)-5]+".wav"
    subprocess.call(['ffmpeg', '-i', s,'-b:a','200k','-map','a',out])
    return speechToText(out)



def handle_rec_file(f):
    g= 'app/static/upload/rec/'+f.name
    with open(g, 'wb+') as destination:  
        for chunk in f.chunks(): 
            destination.write(chunk)
    ind=g.index('.')
    if g[ind:]!='.wav':
        return converterOfFile(g)
    else:
        return speechToText(g)

def calcFeedBack(t):
    keywords=['uem',"university","package",'visit','jaipur','engineering','management','lpa','best','college','rajasthan','placement']
    text=t.split(" ")
    total,acc=len(text),0
    for i in text:
        if i in keywords:
            acc=acc+1
    feed=int((acc/total)*100)
    return feed